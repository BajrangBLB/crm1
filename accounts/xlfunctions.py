import openpyxl
from openpyxl.utils import *
from openpyxl import*
from django.contrib import messages


from .models import *

import re


#load it dynamically the path
wb = openpyxl.load_workbook("media/all_reference_tables.xlsx")
  
ws1 = wb["Sheet1"]
rows = ws1.iter_rows(min_row=2,min_col=1,max_col=2)

GA1=[]

LOa=[]
LOb=[]
LOc=[]
LOd=[]
LOe=[]
LOf=[]

for a,b in rows:
    if b.value == "1a":
        LOa.append(a.value)

    elif b.value == "1b":
        LOb.append(a.value)

    elif b.value == "1c":
        LOc.append(a.value)

    elif b.value == "1d":
        LOd.append(a.value)

    elif b.value == "1e":
        LOe.append(a.value)

    elif b.value == "1f":
        LOf.append(a.value)

GA1.append(LOa)
GA1.append(LOb)
GA1.append(LOc)
GA1.append(LOd)
GA1.append(LOe)
GA1.append(LOf)


ws2 = wb["Sheet2"]
rows = ws2.iter_rows(min_row=2,min_col=1,max_col=2)

GA2=[]

LOa_2=[]
LOb_2=[]
LOc_2=[]
LOd_2=[]
LOe_2=[]
LOf_2=[]
LOg_2=[]

for a,b in rows:
    if b.value == "2a":
        LOa_2.append(a.value)

    elif b.value == "2b":
        LOb_2.append(a.value)

    elif b.value == "2c":
        LOc_2.append(a.value)

    elif b.value == "2d":
        LOd_2.append(a.value)

    elif b.value == "2e":
        LOe_2.append(a.value)

    elif b.value == "2f":
        LOf_2.append(a.value)

    elif b.value == "2g":
        LOg_2.append(a.value)


GA2.append(LOa_2)
GA2.append(LOb_2)
GA2.append(LOc_2)
GA2.append(LOd_2)
GA2.append(LOe_2)
GA2.append(LOf_2)
GA2.append(LOg_2)


 
ws3 = wb["Sheet3"]
rows = ws3.iter_rows(min_row=2,min_col=1,max_col=2)

GA3=[]

LOa_3=[]
LOb_3=[]
LOc_3=[]
LOd_3=[]
LOe_3=[]
LOf_3=[]

for a,b in rows:
    if b.value == "3a":
        LOa_3.append(a.value)

    elif b.value == "3b":
        LOb_3.append(a.value)

    elif b.value == "3c":
        LOc_3.append(a.value)

    elif b.value == "3d":
        LOd_3.append(a.value)

    elif b.value == "3e":
        LOe_3.append(a.value)

    elif b.value == "3f":
        LOf_3.append(a.value)

GA3.append(LOa_3)
GA3.append(LOb_3)
GA3.append(LOc_3)
GA3.append(LOd_3)
GA3.append(LOe_3)
GA3.append(LOf_3)



ws4 = wb["Sheet4"]
rows = ws4.iter_rows(min_row=2,min_col=1,max_col=2)

GA4=[]

LOa_4=[]
LOb_4=[]
LOc_4=[]
LOd_4=[]
LOe_4=[]
LOf_4=[]
LOg_4=[]
LOh_4=[]

for a,b in rows:
    if b.value == "4a":
        LOa_4.append(a.value)

    elif b.value == "4b":
        LOb_4.append(a.value)

    elif b.value == "4c":
        LOc_4.append(a.value)

    elif b.value == "4d":
        LOd_4.append(a.value)

    elif b.value == "4e":
        LOe_4.append(a.value)

    elif b.value == "4f":
        LOf_4.append(a.value)
        
    elif b.value == "4g":
        LOg_4.append(a.value)

    elif b.value == "4h":
        LOh_4.append(a.value)

GA4.append(LOa_4)
GA4.append(LOb_4)
GA4.append(LOc_4)
GA4.append(LOd_4)
GA4.append(LOe_4)
GA4.append(LOf_4)
GA4.append(LOg_4)
GA4.append(LOh_4)


ws5 = wb["Sheet5"]
rows = ws5.iter_rows(min_row=2,min_col=1,max_col=2)

GA5=[]
LOa_5=[]
LOb_5=[]
LOc_5=[]
LOd_5=[]
LOe_5=[]
 
for a,b in rows:
    if b.value == "5a":
        LOa_5.append(a.value)

    elif b.value == "5b":
        LOb_5.append(a.value)

    elif b.value == "5c":
        LOc_5.append(a.value)

    elif b.value == "5d":
        LOd_5.append(a.value)

    elif b.value == "5e":
        LOe_5.append(a.value)

GA5.append(LOa_5)
GA5.append(LOb_5)
GA5.append(LOc_5)
GA5.append(LOd_5)
GA5.append(LOe_5)

    

ws6 = wb["Sheet6"]
rows = ws6.iter_rows(min_row=2,min_col=1,max_col=2)
GA6=[]

LOa_6=[]
LOb_6=[]
LOc_6=[]
LOd_6=[]
LOe_6=[]
LOf_6=[]

for a,b in rows:
    if b.value == "6a":
        LOa_6.append(a.value)

    elif b.value == "6b":
        LOb_6.append(a.value)

    elif b.value == "6c":
        LOc_6.append(a.value)

    elif b.value == "6d":
        LOd_6.append(a.value)

    elif b.value == "6e":
        LOe_6.append(a.value)

    elif b.value == "6f":
        LOf_6.append(a.value)

GA6.append(LOa_6)
GA6.append(LOb_6)
GA6.append(LOc_6)
GA6.append(LOd_6)
GA6.append(LOe_6)
GA6.append(LOf_6)


ws7 = wb["Sheet7"]
rows = ws7.iter_rows(min_row=2,min_col=1,max_col=2)

GA7=[]

LOa_7=[]
LOb_7=[]
LOc_7=[]
LOd_7=[]
LOe_7=[]
LOf_7=[]
LOg_7=[]
LOh_7=[]

for a,b in rows:
    if b.value == "7a":
        LOa_7.append(a.value)

    elif b.value == "7b":
        LOb_7.append(a.value)

    elif b.value == "7c":
        LOc_7.append(a.value)

    elif b.value == "7d":
        LOd_7.append(a.value)

    elif b.value == "7e":
        LOe_7.append(a.value)

    elif b.value == "7f":
        LOf_7.append(a.value)

    elif b.value == "7g":
        LOg_7.append(a.value)
    
    elif b.value == "7h":
        LOh_7.append(a.value)

GA7.append(LOa_7)
GA7.append(LOb_7)
GA7.append(LOc_7)
GA7.append(LOd_7)
GA7.append(LOe_7)
GA7.append(LOf_7)
GA7.append(LOg_7)
GA7.append(LOh_7)

ws8 = wb["Sheet8"]
rows = ws8.iter_rows(min_row=2,min_col=1,max_col=2)

GA8=[]

LOa_8=[]
LOb_8=[]
LOc_8=[]
LOd_8=[]
LOe_8=[]
LOf_8=[]

for a,b in rows:
    if b.value == "8a":
        LOa_8.append(a.value)

    elif b.value == "8b":
        LOb_8.append(a.value)

    elif b.value == "8c":
        LOc_8.append(a.value)

    elif b.value == "8d":
        LOd_8.append(a.value)

    elif b.value == "8e":
        LOe_8.append(a.value)

    elif b.value == "8f":
        LOf_8.append(a.value)


GA8.append(LOa_8)
GA8.append(LOb_8)
GA8.append(LOc_8)
GA8.append(LOd_8)
GA8.append(LOe_8)
GA8.append(LOf_8)
 
ws9 = wb["Sheet9"]
rows = ws9.iter_rows(min_row=2,min_col=1,max_col=2)

GA9=[]

LOa_9=[]
LOb_9=[]
LOc_9=[]
LOd_9=[]
LOe_9=[]
LOf_9=[]
LOg_9=[]

for a,b in rows:
    if b.value == "9a":
        LOa_9.append(a.value)

    elif b.value == "9b":
        LOb_9.append(a.value)

    elif b.value == "9c":
        LOc_9.append(a.value)

    elif b.value == "9d":
        LOd_9.append(a.value)

    elif b.value == "9e":
        LOe_9.append(a.value)

    elif b.value == "9f":
        LOf_9.append(a.value)

    elif b.value == "9g":
        LOg_9.append(a.value)

GA9.append(LOa_9)
GA9.append(LOb_9)
GA9.append(LOc_9)
GA9.append(LOd_9)
GA9.append(LOe_9)
GA9.append(LOf_9)
GA9.append(LOg_9)
 

ws10 = wb["Sheet10"]
rows = ws10.iter_rows(min_row=2,min_col=1,max_col=2)

GA10=[]

LOa_10=[]
LOb_10=[]
LOc_10=[]
LOd_10=[]
LOe_10=[]
LOf_10=[]
 

for a,b in rows:
    if b.value == "10a":
        LOa_10.append(a.value)

    elif b.value == "10b":
        LOb_10.append(a.value)

    elif b.value == "10c":
        LOc_10.append(a.value)

    elif b.value == "10d":
        LOd_10.append(a.value)

    elif b.value == "10e":
        LOe_10.append(a.value)

    elif b.value == "10f":
        LOf_10.append(a.value)

GA10.append(LOa_10)
GA10.append(LOb_10)
GA10.append(LOc_10)
GA10.append(LOd_10)
GA10.append(LOe_10)
GA10.append(LOf_10)


ws11 = wb["Sheet11"]
rows = ws11.iter_rows(min_row=2,min_col=1,max_col=2)

GA11=[]

LOa_11=[]
LOb_11=[]
LOc_11=[]
LOd_11=[]
LOe_11=[]
LOf_11=[]
 

for a,b in rows:
    if b.value == "11a":
        LOa_11.append(a.value)

    elif b.value == "11b":
        LOb_11.append(a.value)

    elif b.value == "11c":
        LOc_11.append(a.value)

    elif b.value == "11d":
        LOd_11.append(a.value)

    elif b.value == "11e":
        LOe_11.append(a.value)

    elif b.value == "11f":
        LOf_11.append(a.value)

GA11.append(LOa_11)
GA11.append(LOb_11)
GA11.append(LOc_11)
GA11.append(LOd_11)
GA11.append(LOe_11)
GA11.append(LOf_11)

ws12 = wb["Sheet12"]
rows = ws12.iter_rows(min_row=2,min_col=1,max_col=2)

GA12=[]

LOa_12=[]
LOb_12=[]
LOc_12=[]
LOd_12=[]
LOe_12=[]
LOf_12=[]
 

for a,b in rows:
    if b.value == "12a":
        LOa_12.append(a.value)

    elif b.value == "12b":
        LOb_12.append(a.value)

    elif b.value == "12c":
        LOc_12.append(a.value)

    elif b.value == "12d":
        LOd_12.append(a.value)

    elif b.value == "12e":
        LOe_12.append(a.value)

    elif b.value == "12f":
        LOf_12.append(a.value)


GA12.append(LOa_12)
GA12.append(LOb_12)
GA12.append(LOc_12)
GA12.append(LOd_12)
GA12.append(LOe_12)
GA12.append(LOf_12)


def getLOname(course):

 
    
        course_name= re.split('-', course)
        course_name1=course_name[0]+"-"+course_name[1]
    
        upfile=(LO_Folder.objects.get(f_name__icontains=course_name1))
        filepath=str(upfile.file_url())

        wb = openpyxl.load_workbook(filepath[1:])

        ws = wb.worksheets[0] 
    

        for column in ws.iter_cols(max_col=1):
            for cell in column:
                if cell.value == course:
                    assessment_methodology= ws.cell(row=cell.row, column=3).value
                    
                    assessment_name= assessment_methodology[0].upper()

                    return assessment_methodology

        wb.close()

     
    



def getCLOlist(course):

    course_name= re.split('-', course)
    course_name1=course_name[0]+"-"+course_name[1]

    upfile=(CLO_Folder.objects.get(f_name__icontains=course_name1))
    filepath=str(upfile.file_url())
    
    #print("this is the file"+ filepath)
    wb2 = openpyxl.load_workbook(filepath[1:])
    ws2 = wb2.worksheets[0]

     
    

    for row in ws2.iter_rows(max_row=6):
        for cell in row:
            if cell.value == course:
                columnletter = get_column_letter(cell.column)
                
                startindex= columnletter+'6'
                #should end more , have till 40
                endindex= columnletter+'44'

                global marksclo
                marksclo = []
                total_marks=ws2.cell(row=5, column=cell.column).value
                #print(total_marks)
     
                range1= ws2[startindex: endindex ]
                for x in range1:
                    for cell in x:
                       #print(cell.value)
                       if cell.value is not None :
                        if isinstance(cell.value,(int,float)):
                                 
                            value = (cell.value/float(total_marks))
                            marksclo.append(value )

                       else:
                            
                            marksclo.append(cell.value )
                return marksclo

        

    wb2.close()
    



def put_course(GA,course_name,GA_subattribute):

    file_name = "media/GA__-_Template.xlsx"
    split_file = file_name.split('__')

    file_path= split_file[0]+"_"+str(GA)+"_"+ split_file[1]

    
    wb = openpyxl.load_workbook(file_path)
    ws = wb[GA_subattribute]
   


    cloList= getCLOlist(course_name)
    LOname = getLOname(course_name)



    for row in ws.iter_rows(max_row=2):
        for cell in row:
            if cell.value == course_name:

                ws.cell(row=2 , column=cell.column).value= LOname

                #save it dynamically
                wb.save(file_path)

               
                row_number = 3
                your_column = cell.column


# need to change this to be universal instead of loa 
            #for x in LOa:
                for i, value in enumerate(cloList, start=row_number):
                    ws.cell(row=i, column=your_column).value = value

                wb.save(file_path)
        wb.close()



    #############################






